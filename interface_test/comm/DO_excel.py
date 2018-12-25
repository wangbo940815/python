import openpyxl,os,json
from openpyxl import load_workbook
from xlrd import book
from comm import constans
from comm.read_conf import *
import re
from comm.connect_db import Connect_db
class Case:
    def __init__(self):
        self.id=None
        self.url=None
        self.method=None
        self.title=None
        self.test_data=None
        self.expect=None
        self.actual=None
        self.book_name=None
        self.sheet_name=None    
class  Do_excel:
    def __init__(self):
        self.excel_path=constans.data_path
    def read_excel(self,sheet_name):
        wb=load_workbook(self.excel_path)
        sheet=wb[sheet_name]
        rows=sheet.max_row
        cases=[]
        url=Read_conf().get("http_request", "url_pre")
        for i in  range(2,rows+1):
            case=Case()
            case.id=sheet.cell(i,1).value
            case.url=url+sheet.cell(i,2).value
            case.method=sheet.cell(i,3).value
            case.title=sheet.cell(i,4).value
            case.test_data=sheet.cell(i,5).value
            case.test_data=case.test_data
            case.expect=   json.loads(sheet.cell(i,6).value)
            case.actual=sheet.cell(i,7).value
#             case.book_name=book_name
            case.sheet_name=sheet_name
            cases.append(case)
        return cases

    def write_excel(self,sheet_name,row,col,value):
        wb=load_workbook(self.excel_path)
        sheet=wb[sheet_name]        
        sheet.cell(row,col).value=value
        wb.save(self.excel_path)
if __name__ == '__main__':
    a='{"mobilephone":"${register_phoneIsRight}","pwd":"1234567890123456789"}'
  