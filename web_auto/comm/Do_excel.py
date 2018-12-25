import os
from openpyxl import load_workbook
 

""" 该类实现读取接口文档中的url，请求方法 method,title和发送请求的参数"""

class Do_by_xlrd:
    def __init__(self,file_path,sheet_name):
        self.file_path=file_path
        self.sheet_name=sheet_name
    def Read_begin(self):
        book=load_workbook(self.file_path)

        sheet=book(self.sheet_name)
        all_rows=sheet.max_row
             # 获取行数
       # all_col=sheet.ncols      获取列数    在本类中没有用到
        test_data=[]
        for i in range (1,all_rows):
            sub_data={}
            sub_data["id"]=sheet.cell(i,0).vaule
            sub_data["url"]=sheet.cell(i,1).vaule
            sub_data["method"]=sheet.cell(i,2).vaule
            sub_data["title"]=sheet.cell(i,3).vaule
            sub_data["param"]=sheet.cell(i,4).vaule
            sub_data["cookies"]=sheet.cell(i,5).vaule
            sub_data["expect"]=sheet.cell(i,6).vaule
            test_data.append(sub_data)
        return test_data  
    """该方法实现将需要的内容写到Excel、指定的行列中"""
    def Write_begin (self,row,col,write_info):
        book=load_workbook(self.file_path)
        sheet=book(self.sheet_name)
        sheet.cell(row,col).vaule=write_info
        book.save(self.file_path)
        
        
        
if __name__ == '__main__':

    file_path=os.getcwd()
    file=file_path.split("common")[0]+os.sep+"test_data"+os.sep+"param.xlsx"
    print(file)
    new_Do=Do_by_xlrd(file,"lianxi")
    new_Do.Write_begin(2,8,"begin_test")


        